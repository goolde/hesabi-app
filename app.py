#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import sqlite3, json, os, hashlib, secrets, io, re
from datetime import datetime, date, timedelta
from functools import wraps
from flask import Flask, request, jsonify, session, send_file, g
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_RIGHT, TA_CENTER

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))

# ============================================================
# دعم PostgreSQL و SQLite تلقائياً
# ============================================================
DATABASE_URL = os.environ.get('DATABASE_URL', '')
USE_POSTGRES = DATABASE_URL.startswith('postgres')

if USE_POSTGRES:
    import psycopg2
    import psycopg2.extras
    print("🐘 استخدام PostgreSQL")
else:
    print("📁 استخدام SQLite")

SMS_PROVIDER = os.environ.get('SMS_PROVIDER', 'demo')
SMS_API_KEY  = os.environ.get('SMS_API_KEY', '')
SMS_SENDER   = os.environ.get('SMS_SENDER', 'Hesabi')
DEVELOPER_PHONE    = '0567867414'
DEVELOPER_NAME     = 'وجيه علي'
DEVELOPER_WHATSAPP = 'https://wa.me/966567867414'

PLANS = {
    'free':         {'name':'مجانية',   'price':0,  'contacts':10, 'transactions':50, 'export':False},
    'professional': {'name':'احترافية', 'price':29, 'contacts':0,  'transactions':0,  'export':True},
    'business':     {'name':'أعمال',    'price':79, 'contacts':0,  'transactions':0,  'export':True},
}

MOYASAR_API_KEY    = os.environ.get('MOYASAR_API_KEY', '')
MOYASAR_SECRET_KEY = os.environ.get('MOYASAR_SECRET_KEY', '')
APP_URL            = os.environ.get('APP_URL', 'http://localhost:5000')


COUNTRY_CURRENCY = {
    'SA': {'name':'السعودية', 'phone_code':'+966', 'currency':'SAR', 'symbol':'ر.س', 'local_regex': r'^(?:0?5\d{8})$'},
    'YE': {'name':'اليمن', 'phone_code':'+967', 'currency':'YER', 'symbol':'ر.ي', 'local_regex': r'^(?:0?[137]\d{7,8})$'},
    'AE': {'name':'الإمارات', 'phone_code':'+971', 'currency':'AED', 'symbol':'د.إ', 'local_regex': r'^(?:0?5\d{8})$'},
    'EG': {'name':'مصر', 'phone_code':'+20',  'currency':'EGP', 'symbol':'ج.م', 'local_regex': r'^(?:0?1\d{9})$'},
    'KW': {'name':'الكويت', 'phone_code':'+965', 'currency':'KWD', 'symbol':'د.ك', 'local_regex': r'^(?:[569]\d{7})$'},
    'QA': {'name':'قطر', 'phone_code':'+974', 'currency':'QAR', 'symbol':'ر.ق', 'local_regex': r'^(?:[3567]\d{7})$'},
    'BH': {'name':'البحرين', 'phone_code':'+973', 'currency':'BHD', 'symbol':'د.ب', 'local_regex': r'^(?:[36]\d{7})$'},
    'OM': {'name':'عمان', 'phone_code':'+968', 'currency':'OMR', 'symbol':'ر.ع', 'local_regex': r'^(?:[79]\d{7})$'},
    'US': {'name':'أمريكا', 'phone_code':'+1',   'currency':'USD', 'symbol':'$',   'local_regex': r'^(?:\d{10})$'},
    'GB': {'name':'بريطانيا', 'phone_code':'+44',  'currency':'GBP', 'symbol':'£',   'local_regex': r'^(?:0?7\d{9})$'},
}

PHONE_CODE_TO_COUNTRY = {v['phone_code']: k for k, v in COUNTRY_CURRENCY.items()}

def currency_symbol(code):
    code = (code or 'SAR').upper()
    for v in COUNTRY_CURRENCY.values():
        if v['currency'] == code:
            return v['symbol']
    return code

def normalize_phone(phone, phone_code='+966'):
    """توحيد رقم الجوال مع رمز البلد والتحقق من صحة الرقم المحلي."""
    raw = (phone or '').strip().replace(' ', '').replace('-', '')
    code = (phone_code or '+966').strip().replace(' ', '')
    if not code.startswith('+'):
        code = '+' + code.lstrip('0')
    if code not in PHONE_CODE_TO_COUNTRY:
        return None, None, 'رمز البلد غير مدعوم'
    country = PHONE_CODE_TO_COUNTRY[code]
    info = COUNTRY_CURRENCY[country]
    if raw.startswith('+'):
        if not raw.startswith(code):
            return None, country, 'رمز البلد لا يطابق رقم الجوال'
        local = raw[len(code):]
    elif raw.startswith('00'):
        full = '+' + raw[2:]
        if not full.startswith(code):
            return None, country, 'رمز البلد لا يطابق رقم الجوال'
        local = full[len(code):]
    else:
        local = raw
    if not re.match(info['local_regex'], local):
        return None, country, f'رقم الجوال غير صحيح لدولة {info["name"]}'
    local = local[1:] if local.startswith('0') else local
    return code + local, country, None

def phone_candidates_for_login(phone):
    raw = (phone or '').strip().replace(' ', '').replace('-', '')
    candidates = [raw]
    for code in PHONE_CODE_TO_COUNTRY:
        n, _, _ = normalize_phone(raw, code)
        if n:
            candidates.append(n)
    if re.match(r'^05\d{8}$', raw):
        candidates.append('+966' + raw[1:])
    return list(dict.fromkeys([x for x in candidates if x]))


# ============================================================
# قاعدة البيانات — يدعم SQLite وPostgreSQL
# ============================================================
def get_db():
    if 'db' not in g:
        if USE_POSTGRES:
            # تصحيح رابط postgres:// لـ postgresql://
            url = DATABASE_URL
            if url.startswith('postgres://'):
                url = url.replace('postgres://', 'postgresql://', 1)
            g.db = psycopg2.connect(url, cursor_factory=psycopg2.extras.RealDictCursor)
            g.db.autocommit = False
        else:
            g.db = sqlite3.connect('hesabi.db')
            g.db.row_factory = sqlite3.Row
            g.db.execute("PRAGMA foreign_keys = ON")
            g.db.execute("PRAGMA journal_mode = WAL")
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop('db', None)
    if db:
        try: db.close()
        except: pass

def db_execute(query, params=(), fetchone=False, fetchall=False, commit=False):
    """دالة موحّدة للاستعلامات — تعمل مع SQLite وPostgreSQL"""
    db = get_db()
    # تحويل ? إلى %s لـ PostgreSQL
    if USE_POSTGRES:
        query = query.replace('?', '%s')
        # تحويل INTEGER PRIMARY KEY AUTOINCREMENT
        query = query.replace('INTEGER PRIMARY KEY AUTOINCREMENT', 'SERIAL PRIMARY KEY')
        query = query.replace('AUTOINCREMENT', '')
    
    if USE_POSTGRES:
        cur = db.cursor()
        cur.execute(query, params)
        if commit:
            db.commit()
        if fetchone:
            return cur.fetchone()
        if fetchall:
            return cur.fetchall()
        return cur
    else:
        cur = db.execute(query, params)
        if commit:
            db.commit()
        if fetchone:
            return cur.fetchone()
        if fetchall:
            return cur.fetchall()
        return cur

def db_executescript(script):
    """تنفيذ script كامل"""
    if USE_POSTGRES:
        db = get_db()
        cur = db.cursor()
        # تقسيم الـ script وتنفيذ كل جملة
        statements = [s.strip() for s in script.split(';') if s.strip()]
        for stmt in statements:
            if stmt:
                stmt = stmt.replace('INTEGER PRIMARY KEY AUTOINCREMENT', 'SERIAL PRIMARY KEY')
                stmt = stmt.replace('AUTOINCREMENT', '')
                stmt = stmt.replace("datetime('now')", 'NOW()')
                stmt = stmt.replace('IF NOT EXISTS', 'IF NOT EXISTS')
                try:
                    cur.execute(stmt)
                except Exception as e:
                    if 'already exists' not in str(e).lower():
                        print(f"Warning: {e}")
        db.commit()
    else:
        db = get_db()
        db.executescript(script)
        db.commit()

def row_to_dict(row):
    """تحويل صف قاعدة البيانات لـ dict"""
    if row is None:
        return None
    if USE_POSTGRES:
        return dict(row)
    return dict(row)

def rows_to_list(rows):
    """تحويل قائمة صفوف لـ list of dicts"""
    if rows is None:
        return []
    return [dict(r) for r in rows]

def init_db():
    """إنشاء الجداول"""
    if USE_POSTGRES:
        db = get_db()
        cur = db.cursor()
        tables = [
            """CREATE TABLE IF NOT EXISTS users (
                id            SERIAL PRIMARY KEY,
                name          TEXT    NOT NULL,
                phone         TEXT    UNIQUE NOT NULL,
                email         TEXT    DEFAULT '',
                business      TEXT    DEFAULT '',
                country       TEXT    DEFAULT 'SA',
                default_currency TEXT DEFAULT 'SAR',
                currencies    TEXT    DEFAULT 'SAR',
                password      TEXT    NOT NULL,
                plan          TEXT    DEFAULT 'free',
                plan_expires  TEXT    DEFAULT '',
                is_active     INTEGER DEFAULT 1,
                otp_code      TEXT    DEFAULT '',
                otp_expires   TEXT    DEFAULT '',
                created       TEXT    DEFAULT NOW()::TEXT,
                last_login    TEXT    DEFAULT '',
                reset_token   TEXT    DEFAULT '',
                reset_expires TEXT    DEFAULT '',
                security_score INTEGER DEFAULT 0
            )""",
            """CREATE TABLE IF NOT EXISTS contacts (
                id            SERIAL PRIMARY KEY,
                user_id       INTEGER NOT NULL,
                name          TEXT    NOT NULL,
                phone         TEXT    DEFAULT '',
                email         TEXT    DEFAULT '',
                type          TEXT    DEFAULT 'عميل',
                notes         TEXT    DEFAULT '',
                lat           REAL,
                lng           REAL,
                location_text TEXT    DEFAULT '',
                created       TEXT    DEFAULT NOW()::TEXT
            )""",
            """CREATE TABLE IF NOT EXISTS transactions (
                id            SERIAL PRIMARY KEY,
                user_id       INTEGER NOT NULL,
                contact_id    INTEGER NOT NULL,
                contact_name  TEXT    DEFAULT '',
                amount        REAL    NOT NULL,
                currency      TEXT    DEFAULT 'ر.س',
                type          TEXT    NOT NULL,
                date          TEXT    DEFAULT '',
                due_date      TEXT    DEFAULT '',
                notes         TEXT    DEFAULT '',
                status        TEXT    DEFAULT 'pending',
                reminder      TEXT    DEFAULT 'none',
                created       TEXT    DEFAULT NOW()::TEXT
            )""",
            """CREATE TABLE IF NOT EXISTS notifications (
                id       SERIAL PRIMARY KEY,
                user_id  INTEGER NOT NULL,
                title    TEXT    DEFAULT '',
                body     TEXT    DEFAULT '',
                icon     TEXT    DEFAULT '📌',
                is_read  INTEGER DEFAULT 0,
                created  TEXT    DEFAULT NOW()::TEXT
            )""",
            """CREATE TABLE IF NOT EXISTS settings (
                user_id  INTEGER PRIMARY KEY,
                data     TEXT    DEFAULT '{}'
            )""",
            """CREATE TABLE IF NOT EXISTS payments (
                id       SERIAL PRIMARY KEY,
                user_id  INTEGER NOT NULL,
                plan     TEXT,
                amount   REAL,
                currency TEXT    DEFAULT 'SAR',
                status   TEXT    DEFAULT 'pending',
                ref_id   TEXT    DEFAULT '',
                created  TEXT    DEFAULT NOW()::TEXT
            )""",
        ]
        for t in tables:
            try:
                cur.execute(t)
            except Exception as e:
                print(f"Table warning: {e}")
        # ترقية الجداول القديمة بدون حذف البيانات
        for col, ddl in {
            'country': "ALTER TABLE users ADD COLUMN country TEXT DEFAULT 'SA'",
            'default_currency': "ALTER TABLE users ADD COLUMN default_currency TEXT DEFAULT 'SAR'",
            'currencies': "ALTER TABLE users ADD COLUMN currencies TEXT DEFAULT 'SAR'",
            'reset_token': "ALTER TABLE users ADD COLUMN reset_token TEXT DEFAULT ''",
            'reset_expires': "ALTER TABLE users ADD COLUMN reset_expires TEXT DEFAULT ''",
            'security_score': "ALTER TABLE users ADD COLUMN security_score INTEGER DEFAULT 0",
        }.items():
            try:
                cur.execute(ddl)
            except Exception as e:
                if 'already exists' not in str(e).lower() and 'duplicate column' not in str(e).lower():
                    print(f"Column warning {col}: {e}")
        db.commit()
    else:
        db = sqlite3.connect('hesabi.db')
        db.executescript("""
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL,
                phone TEXT UNIQUE NOT NULL, email TEXT DEFAULT '',
                business TEXT DEFAULT '', country TEXT DEFAULT 'SA',
                default_currency TEXT DEFAULT 'SAR', currencies TEXT DEFAULT 'SAR',
                password TEXT NOT NULL,
                plan TEXT DEFAULT 'free', plan_expires TEXT DEFAULT '',
                is_active INTEGER DEFAULT 1, otp_code TEXT DEFAULT '',
                otp_expires TEXT DEFAULT '', created TEXT DEFAULT (datetime('now')),
                last_login TEXT DEFAULT '',
                reset_token TEXT DEFAULT '', reset_expires TEXT DEFAULT '', security_score INTEGER DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS contacts (
                id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER NOT NULL,
                name TEXT NOT NULL, phone TEXT DEFAULT '', email TEXT DEFAULT '',
                type TEXT DEFAULT 'عميل', notes TEXT DEFAULT '',
                lat REAL, lng REAL, location_text TEXT DEFAULT '',
                created TEXT DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS transactions (
                id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER NOT NULL,
                contact_id INTEGER NOT NULL, contact_name TEXT DEFAULT '',
                amount REAL NOT NULL, currency TEXT DEFAULT 'ر.س',
                type TEXT NOT NULL, date TEXT DEFAULT '', due_date TEXT DEFAULT '',
                notes TEXT DEFAULT '', status TEXT DEFAULT 'pending',
                reminder TEXT DEFAULT 'none', created TEXT DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS notifications (
                id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER NOT NULL,
                title TEXT DEFAULT '', body TEXT DEFAULT '', icon TEXT DEFAULT '📌',
                is_read INTEGER DEFAULT 0, created TEXT DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS settings (
                user_id INTEGER PRIMARY KEY, data TEXT DEFAULT '{}'
            );
            CREATE TABLE IF NOT EXISTS payments (
                id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER NOT NULL,
                plan TEXT, amount REAL, currency TEXT DEFAULT 'SAR',
                status TEXT DEFAULT 'pending', ref_id TEXT DEFAULT '',
                created TEXT DEFAULT (datetime('now'))
            );
        """)
        # ترقية الجداول القديمة بدون حذف البيانات
        for ddl in [
            "ALTER TABLE users ADD COLUMN country TEXT DEFAULT 'SA'",
            "ALTER TABLE users ADD COLUMN default_currency TEXT DEFAULT 'SAR'",
            "ALTER TABLE users ADD COLUMN currencies TEXT DEFAULT 'SAR'",
            "ALTER TABLE users ADD COLUMN reset_token TEXT DEFAULT ''",
            "ALTER TABLE users ADD COLUMN reset_expires TEXT DEFAULT ''",
            "ALTER TABLE users ADD COLUMN security_score INTEGER DEFAULT 0",
        ]:
            try:
                db.execute(ddl)
            except Exception:
                pass
        db.commit()
        db.close()
    print("✅ قاعدة البيانات جاهزة")

# ============================================================
# دوال مساعدة
# ============================================================
def hash_pwd(p): return hashlib.sha256(p.encode()).hexdigest()
def gen_otp():   return str(secrets.randbelow(9000) + 1000)
def now_str():   return datetime.now().strftime('%Y-%m-%d %H:%M:%S')

def login_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if 'user_id' not in session:
            return jsonify({'error':'غير مسجّل','redirect':'/'}), 401
        return f(*a, **kw)
    return dec

def q(sql, params=(), one=False, many=False, commit=False):
    """اختصار للاستعلامات"""
    db = get_db()
    if USE_POSTGRES:
        sql = sql.replace('?', '%s')
        cur = db.cursor()
        cur.execute(sql, params)
        if commit: db.commit()
        if one:
            r = cur.fetchone()
            return dict(r) if r else None
        if many:
            rows = cur.fetchall()
            return [dict(r) for r in rows]
        return cur
    else:
        cur = db.execute(sql, params)
        if commit: db.commit()
        if one:
            r = cur.fetchone()
            return dict(r) if r else None
        if many:
            rows = cur.fetchall()
            return [dict(r) for r in rows]
        return cur

def commit():
    if USE_POSTGRES:
        get_db().commit()
    # SQLite commits via q()

def check_plan_limit(uid, resource):
    user = q("SELECT plan FROM users WHERE id=?", (uid,), one=True)
    if not user: return False
    plan = PLANS.get(user['plan'], PLANS['free'])
    if resource == 'export': return plan['export']
    limit = plan.get(resource, 0)
    if limit == 0: return True
    count = q(f"SELECT COUNT(*) as c FROM {resource} WHERE user_id=?", (uid,), one=True)
    return (count['c'] if count else 0) < limit

def add_notif(uid, title, body, icon='📌'):
    q("INSERT INTO notifications (user_id,title,body,icon) VALUES (?,?,?,?)",
      (uid,title,body,icon), commit=True)

def get_settings(uid):
    row = q("SELECT data FROM settings WHERE user_id=?", (uid,), one=True)
    if row:
        try: return json.loads(row['data'])
        except: return {}
    return {}

def save_settings(uid, data):
    if USE_POSTGRES:
        q("""INSERT INTO settings (user_id,data) VALUES (%s,%s)
             ON CONFLICT(user_id) DO UPDATE SET data=EXCLUDED.data""",
          (uid, json.dumps(data, ensure_ascii=False)), commit=True)
    else:
        q("INSERT INTO settings (user_id,data) VALUES (?,?) ON CONFLICT(user_id) DO UPDATE SET data=excluded.data",
          (uid, json.dumps(data, ensure_ascii=False)), commit=True)

def send_sms(phone, message):
    if SMS_PROVIDER == 'demo':
        print(f"📱 [SMS Demo] To: {phone} | Msg: {message}")
        return True
    if SMS_PROVIDER == 'unifonic':
        import urllib.request, urllib.parse
        data = urllib.parse.urlencode({'AppSid':SMS_API_KEY,'SenderID':SMS_SENDER,
            'Body':message,'Recipient':phone,'responseType':'JSON'}).encode()
        try: urllib.request.urlopen('https://el.cloud.unifonic.com/rest/SMS/messages',data,timeout=10); return True
        except: return False
    if SMS_PROVIDER == 'taqnyat':
        import urllib.request
        payload = json.dumps({'recipients':[phone],'body':message,'sender':SMS_SENDER}).encode()
        req = urllib.request.Request('https://api.taqnyat.sa/v1/messages',data=payload,
            headers={'Authorization':f'Bearer {SMS_API_KEY}','Content-Type':'application/json'})
        try: urllib.request.urlopen(req,timeout=10); return True
        except: return False
    return False


def send_email(to, subject, message):
    print(f"📧 [Email Demo] To: {to} | Subject: {subject} | Msg: {message}")
    return True

def password_score(pwd):
    score = 0
    if len(pwd or '') >= 8: score += 25
    if re.search(r'[A-Z]', pwd or ''): score += 20
    if re.search(r'[a-z]', pwd or ''): score += 20
    if re.search(r'\d', pwd or ''): score += 20
    if re.search(r'[^A-Za-z0-9]', pwd or ''): score += 15
    return min(score, 100)

def strong_password(pwd):
    return bool(re.match(r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[^A-Za-z0-9]).{8,}$', pwd or ''))

def financial_tip():
    tips = [
        'راجع مستحقاتك اليوم وخفّض المتأخرات خطوة بخطوة.',
        'كل عملية تسجلها اليوم توفر عليك وقت وجهد آخر الشهر.',
        'تابع تاريخ الاستحقاق قبل موعده لتقليل الديون المتأخرة.',
        'افصل معاملاتك حسب الجهة والعملة لتحصل على تقرير أوضح.',
        'التصدير الدوري للبيانات يحميك من ضياع المعلومات.'
    ]
    return tips[secrets.randbelow(len(tips))]

def contact_ids_from_payload(data_or_args):
    if data_or_args is None:
        return []
    if hasattr(data_or_args, 'getlist'):
        vals = data_or_args.getlist('contact_ids') or data_or_args.getlist('contact_id')
    else:
        vals = data_or_args.get('contact_ids') or data_or_args.get('contact_id') or []
    if isinstance(vals, str):
        vals = [x.strip() for x in vals.split(',') if x.strip()]
    if not isinstance(vals, list):
        vals = [vals]
    out=[]
    for v in vals:
        try:
            if str(v).strip(): out.append(int(v))
        except Exception:
            pass
    return list(dict.fromkeys(out))

def apply_statement_filters(sql, params, data):
    if data.get('from'):
        sql += ' AND t.date>=?'; params.append(data['from'])
    if data.get('to'):
        sql += ' AND t.date<=?'; params.append(data['to'])
    ids = contact_ids_from_payload(data)
    if ids:
        placeholders = ','.join(['?'] * len(ids))
        sql += f' AND t.contact_id IN ({placeholders})'
        params.extend(ids)
    if data.get('type'):
        sql += ' AND t.type=?'; params.append(data['type'])
    if data.get('status'):
        sql += ' AND t.status=?'; params.append(data['status'])
    return sql, params

def get_balance(uid, cid):
    rows = q("""SELECT type, SUM(amount) as total FROM transactions
        WHERE user_id=? AND contact_id=? AND status!='settled' GROUP BY type""",
        (uid,cid), many=True)
    fm = om = 0.0
    for r in rows:
        if r['type']=='incoming': fm = r['total'] or 0
        else: om = r['total'] or 0
    return {'for_me':fm,'on_me':om,'net':fm-om}

def load_html():
    html_path = os.path.join(os.path.dirname(__file__), 'templates', 'index.html')
    with open(html_path, 'r', encoding='utf-8') as f:
        return f.read()

# ============================================================
# Routes
# ============================================================
@app.route('/')
def index(): return load_html()

@app.route('/api/register', methods=['POST'])
def register():
    d = request.get_json() or {}
    name = d.get('name','').strip()
    phone = d.get('phone','').strip()
    phone_code = d.get('phone_code', d.get('country_code', '+966')).strip()
    pwd = d.get('password','')
    email = d.get('email','').strip().lower()
    if not all([name, phone, phone_code, email, pwd]):
        return jsonify({'error':'يرجى تعبئة الاسم الكامل ورمز البلد ورقم الجوال والبريد وكلمة المرور'})
    if len(name.split()) < 2:
        return jsonify({'error':'اكتب الاسم كامل على الأقل اسمين'})
    if not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email):
        return jsonify({'error':'البريد الإلكتروني غير صحيح'})
    if not strong_password(pwd):
        return jsonify({'error':'كلمة المرور لازم 8 أحرف وفيها كبير وصغير ورقم ورمز'})
    score = password_score(pwd)
    normalized_phone, country, phone_error = normalize_phone(phone, phone_code)
    if phone_error:
        return jsonify({'error': phone_error})
    default_currency = COUNTRY_CURRENCY[country]['currency']
    currencies = default_currency
    if q("SELECT id FROM users WHERE phone=? OR email=?", (normalized_phone, email), one=True):
        return jsonify({'error':'رقم الجوال أو البريد مسجّل مسبقاً'})
    otp=gen_otp(); expires=(datetime.now()+timedelta(minutes=5)).strftime('%Y-%m-%d %H:%M:%S')
    q("""INSERT INTO users (name,phone,email,country,default_currency,currencies,password,otp_code,otp_expires,security_score)
         VALUES (?,?,?,?,?,?,?,?,?,?)""",
      (name, normalized_phone, email, country, default_currency, currencies, hash_pwd(pwd), otp, expires, score), commit=True)
    send_sms(normalized_phone, f'حسابي: رمز التحقق {otp}. صالح 5 دقائق.')
    print(f"📧 [Email Demo] To: {email} | OTP: {otp}")
    return jsonify({'success':True,'otp_demo':otp,'message':'تم إرسال رمز التحقق للجوال والبريد'})

@app.route('/api/verify-otp', methods=['POST'])
def verify_otp():
    d=request.get_json(); phone=d.get('phone','').strip(); otp=d.get('otp','').strip()
    user = None
    for candidate in phone_candidates_for_login(phone):
        user=q("SELECT * FROM users WHERE phone=?",(candidate,),one=True)
        if user: break
    if not user: return jsonify({'error':'المستخدم غير موجود'})
    if user['otp_code']!=otp: return jsonify({'error':'رمز التحقق غير صحيح'})
    try:
        if datetime.now()>datetime.strptime(user['otp_expires'],'%Y-%m-%d %H:%M:%S'):
            return jsonify({'error':'انتهت صلاحية الرمز'})
    except: pass
    q("UPDATE users SET otp_code='',last_login=? WHERE id=?",(now_str(),user['id']),commit=True)
    session['user_id']=user['id']; session['user_name']=user['name']
    add_notif(user['id'],'مرحباً بك! 👋',f'أهلاً {user["name"]}')
    return jsonify({'success':True,'user':{'id':user['id'],'name':user['name'],'plan':user['plan']}})

@app.route('/api/login', methods=['POST'])
def login():
    d=request.get_json(); phone=d.get('phone','').strip(); pwd=d.get('password','')
    user = None
    hp = hash_pwd(pwd)
    for candidate in phone_candidates_for_login(phone):
        user=q("SELECT * FROM users WHERE phone=? AND password=?",(candidate,hp),one=True)
        if user: break
    if not user: return jsonify({'error':'الجوال أو كلمة المرور غير صحيحة'})
    if not user['is_active']: return jsonify({'error':'الحساب موقوف'})
    q("UPDATE users SET last_login=? WHERE id=?",(now_str(),user['id']),commit=True)
    session['user_id']=user['id']; session['user_name']=user['name']
    tip = financial_tip()
    add_notif(user['id'], 'نصيحة مالية 💡', tip, '💡')
    return jsonify({'success':True,'tip':tip,'user':{'id':user['id'],'name':user['name'],'plan':user['plan']}})

@app.route('/api/logout', methods=['GET', 'POST'])
def logout():
    session.clear()
    return jsonify({'success': True, 'redirect': '/'})

@app.route('/api/me')
def me():
    if 'user_id' not in session: return jsonify({'user':None})
    user=q("SELECT id,name,phone,email,business,country,default_currency,currencies,plan FROM users WHERE id=?",(session['user_id'],),one=True)
    return jsonify({'user':user})

@app.route('/api/resend-otp', methods=['POST'])
def resend_otp():
    d=request.get_json(); phone=d.get('phone','').strip()
    user = None
    for candidate in phone_candidates_for_login(phone):
        user=q("SELECT id FROM users WHERE phone=?",(candidate,),one=True)
        if user: break
    if not user: return jsonify({'error':'المستخدم غير موجود'})
    otp=gen_otp(); expires=(datetime.now()+timedelta(minutes=5)).strftime('%Y-%m-%d %H:%M:%S')
    q("UPDATE users SET otp_code=?,otp_expires=? WHERE id=?",(otp,expires,user['id']),commit=True)
    send_sms(phone,f'حسابي: رمز التحقق {otp}. صالح 5 دقائق.')
    return jsonify({'success':True,'otp_demo':otp})

@app.route('/api/profile', methods=['POST'])
@login_required
def profile():
    d=request.get_json(); uid=session['user_id']
    name=d.get('name','').strip()
    if not name: return jsonify({'error':'الاسم مطلوب'})
    current = q("SELECT country, default_currency, currencies FROM users WHERE id=?", (uid,), one=True) or {}
    country=d.get('country', current.get('country') or 'SA').strip().upper() or 'SA'
    default_currency=d.get('default_currency', current.get('default_currency') or COUNTRY_CURRENCY.get(country, COUNTRY_CURRENCY['SA'])['currency']).strip().upper() or 'SAR'
    currencies=d.get('currencies', current.get('currencies') or default_currency)
    if isinstance(currencies, list): currencies=','.join(currencies)
    cur_list = [x.strip().upper() for x in str(currencies).split(',') if x.strip()]
    if default_currency not in cur_list: cur_list.insert(0, default_currency)
    currencies = ','.join(dict.fromkeys(cur_list))
    q("UPDATE users SET name=?,email=?,business=?,country=?,default_currency=?,currencies=? WHERE id=?",
      (name,d.get('email',''),d.get('business',''),country,default_currency,currencies,uid),commit=True)
    session['user_name']=name
    return jsonify({'success':True})

@app.route('/api/change-password', methods=['POST'])
@login_required
def change_password():
    d=request.get_json(); uid=session['user_id']
    if not q("SELECT id FROM users WHERE id=? AND password=?",(uid,hash_pwd(d.get('current_password',''))),one=True):
        return jsonify({'error':'كلمة المرور الحالية غير صحيحة'})
    new_pwd=d.get('new_password','')
    if not strong_password(new_pwd): return jsonify({'error':'كلمة المرور لازم 8 أحرف وفيها كبير وصغير ورقم ورمز'})
    q("UPDATE users SET password=?, security_score=? WHERE id=?",(hash_pwd(new_pwd),password_score(new_pwd),uid),commit=True)
    return jsonify({'success':True})

@app.route('/api/contacts', methods=['GET','POST'])
@login_required
def contacts():
    uid=session['user_id']
    if request.method=='GET':
        return jsonify({'data':q("SELECT * FROM contacts WHERE user_id=? ORDER BY name",(uid,),many=True)})
    d=request.get_json(); name=d.get('name','').strip()
    if not name: return jsonify({'error':'الاسم مطلوب'})
    if not check_plan_limit(uid,'contacts'):
        return jsonify({'error':'وصلت للحد الأقصى في باقتك','upgrade':True})
    q("INSERT INTO contacts (user_id,name,phone,email,type,notes,lat,lng,location_text) VALUES (?,?,?,?,?,?,?,?,?)",
      (uid,name,d.get('phone',''),d.get('email',''),d.get('type','عميل'),d.get('notes',''),
       d.get('lat'),d.get('lng'),d.get('location_text','')),commit=True)
    add_notif(uid,'جهة جديدة 👤',f'تمت إضافة {name}')
    return jsonify({'success':True})

@app.route('/api/contacts/<int:cid>', methods=['POST'])
@login_required
def contact_op(cid):
    uid=session['user_id']; d=request.get_json()
    if d.get('_method')=='DELETE':
        q("DELETE FROM contacts WHERE id=? AND user_id=?",(cid,uid),commit=True)
        return jsonify({'success':True})
    q("UPDATE contacts SET name=?,phone=?,email=?,type=?,notes=?,lat=?,lng=?,location_text=? WHERE id=? AND user_id=?",
      (d.get('name'),d.get('phone',''),d.get('email',''),d.get('type','عميل'),d.get('notes',''),
       d.get('lat'),d.get('lng'),d.get('location_text',''),cid,uid),commit=True)
    return jsonify({'success':True})

@app.route('/api/contacts/<int:cid>/balance')
@login_required
def contact_balance(cid): return jsonify(get_balance(session['user_id'],cid))

@app.route('/api/transactions', methods=['GET','POST'])
@login_required
def transactions():
    uid=session['user_id']
    if request.method=='GET':
        return jsonify({'data':q("""SELECT t.*,c.name as contact_name FROM transactions t
            LEFT JOIN contacts c ON c.id=t.contact_id
            WHERE t.user_id=? ORDER BY t.created DESC""",(uid,),many=True)})
    d=request.get_json(); cid=d.get('contact_id'); amt=float(d.get('amount',0))
    if not cid or amt<=0: return jsonify({'error':'بيانات غير مكتملة'})
    if not check_plan_limit(uid,'transactions'):
        return jsonify({'error':'وصلت للحد الأقصى في باقتك','upgrade':True})
    c=q("SELECT name FROM contacts WHERE id=?",(cid,),one=True)
    cname=c['name'] if c else '—'
    q("""INSERT INTO transactions (user_id,contact_id,contact_name,amount,currency,type,date,due_date,notes,status,reminder)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
      (uid,cid,cname,amt,d.get('currency','ر.س'),d.get('type','incoming'),
       d.get('date',str(date.today())),d.get('due_date',''),d.get('notes',''),
       d.get('status','pending'),d.get('reminder','none')),commit=True)
    lbl='لي من' if d.get('type')=='incoming' else 'علي من'
    add_notif(uid,'معاملة جديدة 📋',f'{lbl} {cname}: {amt}')
    return jsonify({'success':True})

@app.route('/api/transactions/<int:tid>', methods=['POST'])
@login_required
def tx_op(tid):
    uid=session['user_id']; d=request.get_json()
    if d.get('_method')=='DELETE':
        q("DELETE FROM transactions WHERE id=? AND user_id=?",(tid,uid),commit=True)
        return jsonify({'success':True})
    q("""UPDATE transactions SET contact_id=?,amount=?,currency=?,type=?,date=?,due_date=?,notes=?,status=?
        WHERE id=? AND user_id=?""",
      (d.get('contact_id'),d.get('amount'),d.get('currency','ر.س'),d.get('type'),
       d.get('date'),d.get('due_date',''),d.get('notes',''),d.get('status'),tid,uid),commit=True)
    return jsonify({'success':True})

@app.route('/api/statements', methods=['POST'])
@login_required
def statements():
    uid=session['user_id']; d=request.get_json() or {}
    sql="SELECT t.*,c.name as contact_name,c.type as contact_type FROM transactions t LEFT JOIN contacts c ON c.id=t.contact_id WHERE t.user_id=?"
    p=[uid]
    sql, p = apply_statement_filters(sql, p, d)
    sql += " ORDER BY t.date DESC,t.created DESC"
    return jsonify({'data':q(sql,p,many=True)})

@app.route('/api/notifications')
@login_required
def notifs():
    return jsonify({'data':q("SELECT * FROM notifications WHERE user_id=? ORDER BY created DESC LIMIT 50",
                             (session['user_id'],),many=True)})

@app.route('/api/notifications/read-all', methods=['POST'])
@login_required
def read_notifs():
    q("UPDATE notifications SET is_read=1 WHERE user_id=?",(session['user_id'],),commit=True)
    return jsonify({'success':True})

@app.route('/api/settings', methods=['GET','POST'])
@login_required
def settings():
    uid=session['user_id']
    if request.method=='GET': return jsonify({'data':get_settings(uid)})
    data=get_settings(uid); data.update(request.get_json())
    save_settings(uid,data)
    return jsonify({'success':True})

@app.route('/api/search')
@login_required
def search():
    sq=request.args.get('q','').strip(); uid=session['user_id']
    if len(sq)<2: return jsonify({'contacts':[],'transactions':[]})
    cs=q("SELECT * FROM contacts WHERE user_id=? AND (name LIKE ? OR phone LIKE ?) LIMIT 6",
         (uid,f'%{sq}%',f'%{sq}%'),many=True)
    ts=q("""SELECT t.*,c.name as contact_name FROM transactions t
        LEFT JOIN contacts c ON c.id=t.contact_id
        WHERE t.user_id=? AND (c.name LIKE ? OR t.notes LIKE ?)
        ORDER BY t.created DESC LIMIT 6""",(uid,f'%{sq}%',f'%{sq}%'),many=True)
    return jsonify({'contacts':cs,'transactions':ts})

@app.route('/api/currency-options')
def currency_options():
    return jsonify({'countries': COUNTRY_CURRENCY})

@app.route('/api/plans')
def plans(): return jsonify({'plans':PLANS})

@app.route('/api/my-plan')
@login_required
def my_plan():
    uid=session['user_id']
    user=q("SELECT plan,plan_expires FROM users WHERE id=?",(uid,),one=True)
    plan_info=PLANS.get(user['plan'],PLANS['free'])
    cc=q("SELECT COUNT(*) as c FROM contacts WHERE user_id=?",(uid,),one=True)
    tc=q("SELECT COUNT(*) as c FROM transactions WHERE user_id=?",(uid,),one=True)
    expired=False
    if user['plan_expires']:
        try:
            if datetime.now()>datetime.strptime(user['plan_expires'],'%Y-%m-%d'):
                expired=True
                q("UPDATE users SET plan='free',plan_expires='' WHERE id=?",(uid,),commit=True)
        except: pass
    return jsonify({'plan':user['plan'],'plan_name':plan_info['name'],
                    'plan_expires':user['plan_expires'],'expired':expired,
                    'limits':plan_info,'usage':{'contacts':cc['c'],'transactions':tc['c']}})

@app.route('/api/subscribe', methods=['POST'])
@login_required
def subscribe():
    d=request.get_json(); plan=d.get('plan',''); uid=session['user_id']
    if plan not in PLANS: return jsonify({'error':'باقة غير صحيحة'})
    if PLANS[plan]['price']==0:
        q("UPDATE users SET plan='free',plan_expires='' WHERE id=?",(uid,),commit=True)
        return jsonify({'success':True,'message':'تم التحويل للباقة المجانية'})
    if not MOYASAR_API_KEY:
        return jsonify({'success':False,'coming_soon':True,
                        'message':'بوابة الدفع قيد الإعداد'})
    try:
        import urllib.request, base64
        user=q("SELECT name,phone,email FROM users WHERE id=?",(uid,),one=True)
        payload=json.dumps({'amount':int(PLANS[plan]['price']*100),'currency':'SAR',
            'description':f'حسابي — {PLANS[plan]["name"]}',
            'callback_url':f'{APP_URL}/api/payment/callback',
            'source':{'type':'creditcard','name':user['name'],'company':'hesabi',
                      'number':'','cvc':'','month':'','year':''},
            'metadata':{'user_id':str(uid),'plan':plan}}).encode()
        auth=base64.b64encode(f'{MOYASAR_SECRET_KEY}:'.encode()).decode()
        req=urllib.request.Request('https://api.moyasar.com/v1/payments',data=payload,
            headers={'Content-Type':'application/json','Authorization':f'Basic {auth}'})
        result=json.loads(urllib.request.urlopen(req,timeout=15).read().decode())
        q("INSERT INTO payments (user_id,plan,amount,ref_id,status) VALUES (?,?,?,?,?)",
          (uid,plan,PLANS[plan]['price'],result.get('id',''),'pending'),commit=True)
        return jsonify({'success':True,
                        'payment_url':result.get('source',{}).get('transaction_url',''),
                        'payment_id':result.get('id','')})
    except Exception as e:
        return jsonify({'error':f'خطأ في بوابة الدفع: {str(e)}'})

@app.route('/api/payment/callback', methods=['GET','POST'])
def payment_callback():
    payment_id=request.args.get('id') or (request.get_json() or {}).get('id','')
    if not payment_id:
        return '<h2 style="font-family:Cairo;direction:rtl;text-align:center;color:red">❌ رقم الدفع غير صحيح</h2>'
    try:
        import urllib.request, base64
        auth=base64.b64encode(f'{MOYASAR_SECRET_KEY}:'.encode()).decode()
        req=urllib.request.Request(f'https://api.moyasar.com/v1/payments/{payment_id}',
            headers={'Authorization':f'Basic {auth}'})
        result=json.loads(urllib.request.urlopen(req,timeout=15).read().decode())
        if result.get('status')=='paid':
            meta=result.get('metadata',{}); uid=int(meta.get('user_id',0)); plan=meta.get('plan','')
            if uid and plan in PLANS:
                expires=(datetime.now()+timedelta(days=30)).strftime('%Y-%m-%d')
                q("UPDATE users SET plan=?,plan_expires=? WHERE id=?",(plan,expires,uid),commit=True)
                q("UPDATE payments SET status='paid' WHERE ref_id=?",(payment_id,),commit=True)
                add_notif(uid,'تم تفعيل اشتراكك! 🎉',f'باقة {PLANS[plan]["name"]} حتى {expires}')
                return f'''<html><head><meta charset="utf-8"><meta http-equiv="refresh" content="3;url=/">
                <style>body{{font-family:Cairo,sans-serif;direction:rtl;text-align:center;background:#0a0f1e;color:#e2e8f0;padding:60px;}}</style></head>
                <body><div style="font-size:60px;">🎉</div><h2 style="color:#f6c90e;">تم الدفع بنجاح!</h2>
                <p>جاري التحويل...</p></body></html>'''
    except: pass
    return '''<html><head><meta charset="utf-8"><meta http-equiv="refresh" content="4;url=/">
    <style>body{font-family:Cairo,sans-serif;direction:rtl;text-align:center;background:#0a0f1e;color:#e2e8f0;padding:60px;}</style></head>
    <body><div style="font-size:60px;">❌</div><h2 style="color:#fc8181;">فشل الدفع</h2>
    <p>يرجى المحاولة مرة أخرى</p></body></html>'''

@app.route('/api/payment/history')
@login_required
def payment_history():
    return jsonify({'data':q("SELECT * FROM payments WHERE user_id=? ORDER BY created DESC",
                             (session['user_id'],),many=True)})

@app.route('/api/export/pdf')
@login_required
def export_pdf():
    uid=session['user_id']
    if not check_plan_limit(uid,'export'):
        return jsonify({'error':'التصدير للباقة الاحترافية فقط'}),403
    user=q("SELECT * FROM users WHERE id=?",(uid,),one=True)
    s=get_settings(uid)
    filters = request.args
    sql="SELECT t.*,c.name as contact_name FROM transactions t LEFT JOIN contacts c ON c.id=t.contact_id WHERE t.user_id=?"
    p=[uid]
    data = {'from':filters.get('from',''), 'to':filters.get('to',''), 'contact_ids':filters.getlist('contact_ids') or filters.get('contact_id','')}
    sql, p = apply_statement_filters(sql, p, data)
    rows=q(sql+" ORDER BY t.date DESC",p,many=True)
    gold=colors.HexColor('#f6c90e'); dark=colors.HexColor('#1a2340')
    green_=colors.HexColor('#48bb78'); red_=colors.HexColor('#fc8181')
    buf=io.BytesIO()
    doc=SimpleDocTemplate(buf,pagesize=A4,rightMargin=1.5*cm,leftMargin=1.5*cm,topMargin=2*cm,bottomMargin=2*cm)
    title_s=ParagraphStyle('t',fontName='Helvetica-Bold',fontSize=18,alignment=TA_CENTER,textColor=dark,spaceAfter=8)
    sub_s=ParagraphStyle('s',fontName='Helvetica',fontSize=10,alignment=TA_CENTER,textColor=colors.grey,spaceAfter=16)
    story=[]
    biz=user['business'] or user['name']
    story.append(Paragraph(f'Hesabi | {biz}',title_s))
    story.append(Paragraph(f'Account Statement — {datetime.now().strftime("%Y-%m-%d")}',sub_s))
    story.append(HRFlowable(width="100%",thickness=2,color=gold,spaceAfter=12))
    fm=sum(float(r['amount']) for r in rows if r['type']=='incoming' and r['status']!='settled')
    om=sum(float(r['amount']) for r in rows if r['type']=='outgoing' and r['status']!='settled')
    nt=fm-om; cur=s.get('currency','ر.س')
    summ=Table([['For Me (لي)',f'{fm:,.2f} {cur}'],['On Me (علي)',f'{om:,.2f} {cur}'],
                ['Net (الصافي)',f'{nt:+,.2f} {cur}'],['Total',str(len(rows))]],colWidths=[10*cm,7*cm])
    summ.setStyle(TableStyle([('FONTNAME',(0,0),(-1,-1),'Helvetica'),('FONTSIZE',(0,0),(-1,-1),11),
        ('FONTNAME',(1,0),(1,-1),'Helvetica-Bold'),('TEXTCOLOR',(1,0),(1,0),green_),
        ('TEXTCOLOR',(1,1),(1,1),red_),('TEXTCOLOR',(1,2),(1,2),green_ if nt>=0 else red_),
        ('ROWBACKGROUNDS',(0,0),(-1,-1),[colors.white,colors.HexColor('#f5f5f5')]),
        ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#dddddd')),
        ('TOPPADDING',(0,0),(-1,-1),7),('BOTTOMPADDING',(0,0),(-1,-1),7)]))
    story.append(summ); story.append(Spacer(1,16))
    tdata=[['Date','Contact','Notes','For Me','On Me','Status']]
    for r in rows:
        isIn=r['type']=='incoming'
        tdata.append([r['date'] or '—',r['contact_name'] or '—',(r['notes'] or '—')[:28],
            f"{r['amount']} {r['currency']}" if isIn else '—',
            f"{r['amount']} {r['currency']}" if not isIn else '—',
            {'pending':'Pending','partial':'Partial','settled':'Settled'}.get(r['status'],'')])
    mt=Table(tdata,colWidths=[2.5*cm,3.5*cm,4*cm,3*cm,3*cm,2*cm],repeatRows=1)
    mt.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),dark),('TEXTCOLOR',(0,0),(-1,0),gold),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,0),10),
        ('FONTNAME',(0,1),(-1,-1),'Helvetica'),('FONTSIZE',(0,1),(-1,-1),9),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#f9f9f9')]),
        ('GRID',(0,0),(-1,-1),0.3,colors.HexColor('#e0e0e0')),
        ('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5)]))
    story.append(mt); story.append(Spacer(1,16))
    story.append(HRFlowable(width="100%",thickness=1,color=colors.grey))
    foot=ParagraphStyle('f',fontName='Helvetica',fontSize=9,alignment=TA_CENTER,textColor=colors.grey)
    story.append(Paragraph(f'Hesabi | {user["name"]} | {datetime.now().strftime("%Y-%m-%d")}',foot))
    doc.build(story); buf.seek(0)
    return send_file(buf,mimetype='application/pdf',as_attachment=True,download_name=f'hesabi_{date.today()}.pdf')

@app.route('/api/export/excel')
@login_required
def export_excel():
    uid=session['user_id']
    if not check_plan_limit(uid,'export'):
        return jsonify({'error':'التصدير للباقة الاحترافية فقط'}),403
    user=q("SELECT * FROM users WHERE id=?",(uid,),one=True); s=get_settings(uid)
    filters = request.args
    sql="SELECT t.*,c.name as contact_name FROM transactions t LEFT JOIN contacts c ON c.id=t.contact_id WHERE t.user_id=?"
    p=[uid]
    data = {'from':filters.get('from',''), 'to':filters.get('to',''), 'contact_ids':filters.getlist('contact_ids') or filters.get('contact_id','')}
    sql, p = apply_statement_filters(sql, p, data)
    rows=q(sql+" ORDER BY t.date DESC",p,many=True); cur=s.get('currency','ر.س')
    wb=Workbook(); ws=wb.active; ws.title="Statement"; ws.sheet_view.rightToLeft=True
    gold_f=PatternFill("solid",fgColor="F6C90E"); dark_f=PatternFill("solid",fgColor="1A2340")
    ctr=Alignment(horizontal="center",vertical="center"); rt=Alignment(horizontal="right",vertical="center")
    thin=Border(left=Side(style='thin',color='DDDDDD'),right=Side(style='thin',color='DDDDDD'),
                top=Side(style='thin',color='DDDDDD'),bottom=Side(style='thin',color='DDDDDD'))
    ws.merge_cells('A1:H1'); ws['A1']=f'حسابي — {user["business"] or user["name"]}'
    ws['A1'].font=Font(bold=True,size=16,color="F6C90E"); ws['A1'].fill=dark_f; ws['A1'].alignment=ctr
    ws.row_dimensions[1].height=40
    ws.merge_cells('A2:H2'); ws['A2']=f'تاريخ التصدير: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws['A2'].font=Font(size=10,color="AAAAAA"); ws['A2'].fill=PatternFill("solid",fgColor="0F1628"); ws['A2'].alignment=ctr
    fm=sum(float(r['amount']) for r in rows if r['type']=='incoming' and r['status']!='settled')
    om=sum(float(r['amount']) for r in rows if r['type']=='outgoing' and r['status']!='settled')
    ws.row_dimensions[4].height=30
    for ci,(lbl,val,color) in enumerate([('إجمالي لي',f'{fm:,.2f} {cur}','48BB78'),
        ('إجمالي علي',f'{om:,.2f} {cur}','FC8181'),
        ('الصافي',f'{fm-om:+,.2f} {cur}','48BB78' if fm>=om else 'FC8181'),
        ('المعاملات',str(len(rows)),'F6C90E')],1):
        ws.merge_cells(f'{get_column_letter(ci*2-1)}4:{get_column_letter(ci*2)}4')
        c=ws[f'{get_column_letter(ci*2-1)}4']; c.value=f'{lbl}: {val}'
        c.font=Font(bold=True,size=11,color=color); c.fill=PatternFill("solid",fgColor="151D35"); c.alignment=ctr
    hdrs=['التاريخ','الجهة','النوع','المبلغ','العملة','البيان','الحالة','الاستحقاق']
    for ci,h in enumerate(hdrs,1):
        c=ws.cell(row=6,column=ci,value=h); c.font=Font(bold=True,size=10)
        c.fill=gold_f; c.alignment=ctr; c.border=thin
    ws.row_dimensions[6].height=28
    sl={'pending':'معلقة','partial':'جزئية','settled':'مسوّاة'}; tl={'incoming':'لي 📥','outgoing':'علي 📤'}
    for i,r in enumerate(rows,7):
        isIn=r['type']=='incoming'; fill=PatternFill("solid",fgColor="E8F5E9" if isIn else "FFEBEE")
        vals=[r['date'] or '',r['contact_name'] or '',tl.get(r['type'],r['type']),float(r['amount']),
              r['currency'],r['notes'] or '',sl.get(r['status'],r['status']),r['due_date'] or '']
        for ci,v in enumerate(vals,1):
            c=ws.cell(row=i,column=ci,value=v); c.fill=fill if ci<=4 else PatternFill()
            c.alignment=rt; c.border=thin
            if ci==4: c.font=Font(bold=True,color="1B5E20" if isIn else "B71C1C"); c.number_format='#,##0.00'
        ws.row_dimensions[i].height=22
    for ci,w in enumerate([14,20,12,14,10,28,12,14],1):
        ws.column_dimensions[get_column_letter(ci)].width=w
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,download_name=f'hesabi_{date.today()}.xlsx')


@app.route('/api/forgot-password', methods=['POST'])
def forgot_password():
    d=request.get_json() or {}; login=(d.get('login') or d.get('phone') or d.get('email') or '').strip().lower()
    if not login: return jsonify({'error':'اكتب رقم الجوال أو البريد'})
    user = None
    if '@' in login:
        user=q("SELECT id,phone,email FROM users WHERE lower(email)=?", (login,), one=True)
    else:
        for candidate in phone_candidates_for_login(login):
            user=q("SELECT id,phone,email FROM users WHERE phone=?", (candidate,), one=True)
            if user: break
    if not user: return jsonify({'error':'الحساب غير موجود'})
    token=gen_otp(); expires=(datetime.now()+timedelta(minutes=10)).strftime('%Y-%m-%d %H:%M:%S')
    q("UPDATE users SET reset_token=?, reset_expires=? WHERE id=?", (token, expires, user['id']), commit=True)
    send_sms(user.get('phone',''), f'حسابي: رمز إعادة تعيين كلمة المرور {token}. صالح 10 دقائق.')
    if user.get('email'): send_email(user['email'], 'إعادة تعيين كلمة المرور', f'رمز إعادة التعيين: {token}')
    return jsonify({'success':True,'reset_demo':token,'message':'تم إرسال رمز إعادة التعيين'})

@app.route('/api/reset-password', methods=['POST'])
def reset_password():
    d=request.get_json() or {}; login=(d.get('login') or '').strip().lower(); token=(d.get('token') or '').strip(); new_pwd=d.get('new_password','')
    if not all([login,token,new_pwd]): return jsonify({'error':'أكمل البيانات'})
    if not strong_password(new_pwd): return jsonify({'error':'كلمة المرور لازم 8 أحرف وفيها كبير وصغير ورقم ورمز'})
    user = None
    if '@' in login:
        user=q("SELECT * FROM users WHERE lower(email)=?", (login,), one=True)
    else:
        for candidate in phone_candidates_for_login(login):
            user=q("SELECT * FROM users WHERE phone=?", (candidate,), one=True)
            if user: break
    if not user: return jsonify({'error':'الحساب غير موجود'})
    if user.get('reset_token') != token: return jsonify({'error':'رمز الاستعادة غير صحيح'})
    try:
        if datetime.now() > datetime.strptime(user.get('reset_expires',''), '%Y-%m-%d %H:%M:%S'):
            return jsonify({'error':'انتهت صلاحية الرمز'})
    except Exception: pass
    q("UPDATE users SET password=?, reset_token='', reset_expires='', security_score=? WHERE id=?", (hash_pwd(new_pwd), password_score(new_pwd), user['id']), commit=True)
    return jsonify({'success':True,'message':'تم تغيير كلمة المرور'})

@app.route('/api/backup')
@login_required
def backup_json():
    uid=session['user_id']
    return jsonify({
        'created_at': now_str(),
        'user': q("SELECT id,name,phone,email,business,country,default_currency,currencies,plan FROM users WHERE id=?", (uid,), one=True),
        'settings': get_settings(uid),
        'contacts': q("SELECT * FROM contacts WHERE user_id=? ORDER BY name", (uid,), many=True),
        'transactions': q("SELECT * FROM transactions WHERE user_id=? ORDER BY created DESC", (uid,), many=True),
        'notifications': q("SELECT * FROM notifications WHERE user_id=? ORDER BY created DESC", (uid,), many=True),
    })

@app.route('/api/share-statement')
@login_required
def share_statement():
    base = APP_URL.rstrip('/')
    return jsonify({'success':True,'message':'رابط الكشف جاهز', 'url': base + '/plans'})

@app.route('/plans')
def plans_page():
    uid=session.get('user_id'); cur_plan='free'; user_name=''
    if uid:
        user=q("SELECT plan,name FROM users WHERE id=?",(uid,),one=True)
        if user: cur_plan=user['plan']; user_name=user['name']
    plans_html=''
    for pid,p in PLANS.items():
        is_current=pid==cur_plan
        badge=''
        if pid=='professional': badge='<div style="position:absolute;top:-12px;right:16px;background:#3182ce;color:#fff;font-size:11px;font-weight:700;padding:4px 12px;border-radius:20px;">⭐ مميز</div>'
        if pid=='business': badge='<div style="position:absolute;top:-12px;right:16px;background:#f6c90e;color:#000;font-size:11px;font-weight:700;padding:4px 12px;border-radius:20px;">💎 أعمال</div>'
        feats={'free':['50 معاملة شهرياً','10 جهات اتصال','تقارير أساسية','بدون تصدير'],
               'professional':['معاملات غير محدودة','جهات غير محدودة','تقارير متقدمة','تصدير PDF وExcel','نسخ احتياطي','دعم أولوية'],
               'business':['كل مميزات الاحترافية','حتى 5 مستخدمين','فواتير وإيصالات','مزامنة فورية','مدير حساب مخصص']}
        feats_html=''.join([f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:8px;font-size:13px;"><span style="color:#48bb78;font-weight:700;">✓</span>{f}</div>' for f in feats.get(pid,[])])
        if is_current: btn=f'<div style="background:#1a2340;color:#a0aec0;border-radius:10px;padding:12px;text-align:center;font-weight:700;font-size:14px;">✅ باقتك الحالية</div>'
        else:
            price_label="مجاني — ابدأ الآن" if p["price"]==0 else f'اشترك — {p["price"]} ر.س/شهر'
            btn=f'<button onclick="choosePlan(\'{pid}\')" style="width:100%;padding:13px;background:linear-gradient(135deg,#f6c90e,#d4a017);color:#000;border:none;border-radius:10px;font-size:15px;font-weight:700;font-family:Cairo,sans-serif;cursor:pointer;">{price_label}</button>'
        plans_html+=f'''<div style="border:{("2px solid #f6c90e" if is_current else "1px solid rgba(99,179,237,0.2)")};border-radius:16px;padding:24px;background:#111827;position:relative;margin-bottom:20px;">
            {badge}<div style="font-size:20px;font-weight:800;margin-bottom:8px;">{p["name"]}</div>
            <div style="font-size:32px;font-weight:900;color:#f6c90e;margin-bottom:16px;">{p["price"]}<span style="font-size:14px;color:#a0aec0;font-weight:400;"> ر.س / شهر</span></div>
            {feats_html}<div style="margin-top:16px;">{btn}</div></div>'''
    return f'''<!DOCTYPE html><html lang="ar" dir="rtl">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>حسابي — الباقات</title>
<link href="https://fonts.googleapis.com/css2?family=Cairo:wght@400;600;700;900&display=swap" rel="stylesheet">
<style>*{{margin:0;padding:0;box-sizing:border-box;}}body{{font-family:Cairo,sans-serif;background:#0a0f1e;color:#e2e8f0;min-height:100vh;}}</style>
</head><body>
<div style="background:#111827;border-bottom:1px solid rgba(246,201,14,0.3);padding:16px 20px;display:flex;align-items:center;justify-content:space-between;">
    <div style="display:flex;align-items:center;gap:10px;">
        <div style="width:38px;height:38px;background:linear-gradient(135deg,#f6c90e,#d4a017);border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:18px;">💰</div>
        <span style="font-size:18px;font-weight:900;color:#f6c90e;">حسابي</span>
    </div>
    <a href="/" style="color:#a0aec0;text-decoration:none;font-size:13px;">← العودة للتطبيق</a>
</div>
<div style="max-width:480px;margin:0 auto;padding:30px 20px;">
    <div style="text-align:center;margin-bottom:32px;">
        <div style="font-size:40px;margin-bottom:12px;">💎</div>
        <h1 style="font-size:28px;font-weight:900;color:#f6c90e;margin-bottom:8px;">الباقات والأسعار</h1>
        {f'<p style="color:#48bb78;font-size:13px;margin-top:8px;">مرحباً {user_name} 👋</p>' if user_name else ''}
    </div>
    {plans_html}
    <div style="text-align:center;margin-top:24px;padding:20px;background:#111827;border-radius:16px;border:1px solid rgba(99,179,237,0.15);">
        <div style="font-size:24px;margin-bottom:10px;">👨‍💻</div>
        <div style="font-size:15px;font-weight:700;margin-bottom:14px;">تواصل مع المطوّر</div>
        <a href="https://wa.me/966567867414" target="_blank"
           style="display:inline-block;padding:10px 24px;background:rgba(72,187,120,0.15);color:#48bb78;border:1px solid rgba(72,187,120,0.3);border-radius:10px;text-decoration:none;font-weight:700;font-size:14px;">💬 واتساب</a>
    </div>
</div>
<script>
function choosePlan(plan){{
    if(!{str(bool(uid)).lower()}){{window.location.href='/';return;}}
    if(plan==='free'){{
        if(confirm('التحويل للباقة المجانية؟')){{
            fetch('/api/subscribe',{{method:'POST',headers:{{'Content-Type':'application/json'}},body:JSON.stringify({{plan:'free'}})}})
            .then(r=>r.json()).then(d=>{{if(d.success){{alert('تم!');location.reload();}}else alert(d.message||d.error);}});
        }}
        return;
    }}
    const prices={{professional:29,business:79}};
    if(confirm(`الاشتراك في الباقة\nالسعر: ${{prices[plan]}} ر.س/شهر\n\nمتابعة للدفع؟`)){{
        fetch('/api/subscribe',{{method:'POST',headers:{{'Content-Type':'application/json'}},body:JSON.stringify({{plan:plan}})}})
        .then(r=>r.json()).then(d=>{{
            if(d.payment_url){{window.location.href=d.payment_url;}}
            else if(d.coming_soon){{alert('بوابة الدفع قيد الإعداد\nللتفعيل: 0567867414');}}
            else{{alert(d.error||d.message||'خطأ');}}
        }});
    }}
}}
</script></body></html>'''

# ============================================================
# تشغيل التطبيق
# ============================================================
with app.app_context():
    init_db()

if __name__ == '__main__':
    print("="*55)
    print("  💰 حسابي SaaS")
    print("="*55)
    port=int(os.environ.get('PORT',5000))
    debug=os.environ.get('DEBUG','true').lower()=='true'
    print(f"🌐 http://localhost:{port}")
    print(f"🐘 PostgreSQL: {USE_POSTGRES}")
    print("🛑 Ctrl+C للإيقاف")
    print("="*55)
    app.run(debug=debug,host='0.0.0.0',port=port)